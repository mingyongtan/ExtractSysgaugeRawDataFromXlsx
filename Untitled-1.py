"""
Process List Extractor – Desktop GUI (PySide6) — GUI‑Only
-------------------------------------------------------
Fixes in this version
---------------------
1) Removed tkinter (fixed `ModuleNotFoundError`).
2) GUI‑only: no CLI prompts or flags needed (tests still available via `--run-tests`).
3) Robust Excel/CSV parsing and Process List detection.
4) Correct `pandas.to_csv` arg **`lineterminator`** and **no multi‑line string literals** (fixes unterminated string errors).
5) **Multi‑file TXT export**: pick many input files → writes **one .txt per file** with names you choose (filenames sanitized & de‑duplicated).

What it does
------------
- Accepts **.xlsx/.xls/.xlsm/.csv/.csx** SysGauge reports.
- Auto‑detects the **Process List** (row containing "Process Name").
- Cleans headers, trims empty rows.
- **Reorders and formats** to exactly this schema:
  `Process Name\tInstances\tCPU\tMemory\tThreads\tHandles\tData\tStatus`
  - CPU/Memory/Data → **2 decimals** (e.g., `658.21`).
  - Threads/Handles → **thousands separators** (e.g., `34,504`).
  - Instances → integer.
- **Multi‑file TXT export**: select multiple files and export **one .txt per input file** (you pick each filename).

Quick start – Desktop GUI (recommended)
-------------------------------------
1) Install deps (virtualenv recommended):
   pip install PySide6 pandas openpyxl

2) Run the desktop app:
   python ProcessListExtractor_GUI.py

3) Click **Open Excel/CSV (multiple)**, preview, then **Save .txt files (choose names)**.

Run tests (optional)
--------------------
   python ProcessListExtractor_GUI.py --run-tests
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path
from typing import Optional, List, Tuple

import pandas as pd

SUPPORTED_EXTS = {".xlsx", ".xlsm", ".xls", ".csv", ".csx"}
EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}
CSV_EXTS = {".csv", ".csx"}

# Desired final column order and formatting instructions
TARGET_COLS = [
    "Process Name", "Instances", "CPU", "Memory", "Threads", "Handles", "Data", "Status"
]


# -----------------------------
# Core extraction helpers
# -----------------------------

def _normalize(name: str) -> str:
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _drop_placeholder_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Remove columns named variants of Unnamed or NaN/None
    keep = []
    for c in df.columns:
        s = str(c).strip()
        low = s.lower()
        if s == "" or low.startswith("unnamed") or low in {"nan", "none"}:
            continue
        keep.append(c)
    return df[keep]


def reorder_and_format_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return df restricted to TARGET_COLS in order, with formatted values.
    If some target columns are missing, keep the ones we can find; others are created as 0/blank.
    """
    # Map actual columns to normalized keys
    norm_map = {_normalize(c): c for c in df.columns}

    ordered = {}
    for target in TARGET_COLS:
        key = _normalize(target)
        actual = norm_map.get(key)
        if actual is None:
            # Try loose contains, e.g., "processname" within a longer header
            candidates = [c for k, c in norm_map.items() if key in k]
            actual = candidates[0] if candidates else None
        if actual is not None:
            ordered[target] = df[actual]
        else:
            ordered[target] = pd.Series([None] * len(df))

    out = pd.DataFrame(ordered)

    # Coerce numerics
    for col in ["Instances", "Threads", "Handles"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype("int64")
    for col in ["CPU", "Memory", "Data"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0).astype(float)

    # Format per spec
    out["CPU"] = out["CPU"].map(lambda x: f"{x:.2f}")
    out["Memory"] = out["Memory"].map(lambda x: f"{x:.2f}")
    out["Data"] = out["Data"].map(lambda x: f"{x:.2f}")
    out["Threads"] = out["Threads"].map(lambda x: f"{x:,}")
    out["Handles"] = out["Handles"].map(lambda x: f"{x:,}")
    out["Instances"] = out["Instances"].astype(str)

    # Ensure correct column order
    out = out[TARGET_COLS]
    return out


def _find_header_row(df: pd.DataFrame, header_keyword: str = "Process Name") -> Optional[int]:
    """Return index of the first row that contains `header_keyword`.

    The DataFrame should be read with header=None for robust scanning.
    """
    mask = df.apply(lambda r: r.astype(str).str.contains(header_keyword, case=False, na=False)).any(axis=1)
    idxs = list(df.index[mask])
    return idxs[0] if idxs else None


def _clean_columns(cols: List[str]) -> List[str]:
    cleaned: List[str] = []
    seen = set()
    for c in cols:
        name = str(c).replace("\n", " ").strip() or "Unnamed"
        base = name
        i = 2
        while name in seen:
            name = f"{base}_{i}"
            i += 1
        cleaned.append(name)
        seen.add(name)
    return cleaned


def _postprocess(df: pd.DataFrame) -> pd.DataFrame:
    # Keep rows where first column (Process Name) is not empty; drop all-empty rows
    first_col = df.columns[0]
    df = df.loc[df[first_col].astype(str).str.strip().ne("")]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def extract_process_list_from_filelike(
    file_like: io.BytesIO | io.StringIO | io.BufferedReader,
    ext: str,
    csv_delim: Optional[str] = None,
) -> pd.DataFrame:
    """Extract the Process List table from an **open file-like object**.

    Parameters
    ----------
    file_like : file-like object supporting .read()
    ext       : file extension (e.g., ".xlsx" or ".csv") used to decide parser
    csv_delim : optional delimiter override for CSV/CSX
    """
    ext = ext.lower()

    if ext in EXCEL_EXTS:
        xls = pd.ExcelFile(file_like, engine="openpyxl")
        # Prefer a sheet named exactly "Process List"
        preferred = [s for s in xls.sheet_names if s.strip().lower() == "process list"]
        candidates = preferred + [s for s in xls.sheet_names if s not in preferred]
        for sheet in candidates:
            raw = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
            hdr = _find_header_row(raw)
            if hdr is None:
                continue
            headers = _clean_columns(list(raw.iloc[hdr].tolist()))
            body = raw.iloc[hdr + 1 :].copy()
            body.columns = headers
            # Remove empty/unnamed columns
            keep = [c for c in body.columns if c and not str(c).lower().startswith("unnamed")]
            if keep:
                body = body[keep]
            # Drop placeholder columns then reorder/format
            body = _drop_placeholder_columns(body)
            body = reorder_and_format_columns(body)
            return _postprocess(body)
        raise ValueError("Couldn't locate a 'Process Name' header in any sheet.")

    if ext in CSV_EXTS:
        # Read as raw with no header, try provided delimiter or sniff via python engine
        try:
            raw = pd.read_csv(
                file_like,
                header=None,
                sep=csv_delim or None,
                engine="python",
                dtype=str,
                keep_default_na=False,
            )
        except Exception:
            # Fallback to comma
            try:
                file_like.seek(0)
            except Exception:
                pass
            raw = pd.read_csv(file_like, header=None, sep=",", engine="python", dtype=str, keep_default_na=False)
        hdr = _find_header_row(raw)
        if hdr is None:
            raise ValueError("Couldn't locate a 'Process Name' header in the CSV/CSX file.")
        headers = _clean_columns(list(raw.iloc[hdr].tolist()))
        body = raw.iloc[hdr + 1 :].copy()
        body.columns = headers
        keep = [c for c in body.columns if c and not str(c).lower().startswith("unnamed")]
        if keep:
            body = body[keep]
        body = _drop_placeholder_columns(body)
        body = reorder_and_format_columns(body)
        return _postprocess(body)

    raise ValueError(f"Unsupported file extension: {ext}")


def extract_process_list_from_path(path: Path, csv_delim: Optional[str] = None) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext in EXCEL_EXTS:
        with open(path, "rb") as f:
            return extract_process_list_from_filelike(io.BytesIO(f.read()), ext)
    if ext in CSV_EXTS:
        with open(path, "rb") as f:
            data = f.read()
        return extract_process_list_from_filelike(io.StringIO(data.decode("utf-8", errors="ignore")), ext, csv_delim)
    raise ValueError(f"Unsupported file extension: {ext}")


# -----------------------------
# Helpers for labels/filenames
# -----------------------------

# For display labels (and for tests that simulate sheet names)
_INVALID_SHEET_CHARS = set('[]:*?/\\')

def _sheet_title_from_filename(path: str) -> str:
    """Sanitize the filename stem into a valid Excel-like sheet title (<=31 chars).
    Used for the preview "Source" label and preserved for tests.
    """
    base = Path(path).stem
    name = "".join(ch for ch in base if ch not in _INVALID_SHEET_CHARS).strip() or "Sheet"
    return name[:31]


def _make_unique_sheet_name(proposed: str, used: set[str]) -> str:
    """Ensure names are unique (used in tests and when simulating sheet names)."""
    name = proposed[:31]
    if name not in used:
        used.add(name)
        return name
    i = 2
    while True:
        suffix = f"_{i}"
        trimmed = name[: 31 - len(suffix)]
        candidate = trimmed + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1

# For TXT file outputs (Windows-safe)
_INVALID_FILE_CHARS = set('<>:"/\\|?*')

def _safe_txt_filename(path: str, used_lower: set[str]) -> str:
    """Derive a safe, de‑duplicated TXT filename from an input path.
    - Strips illegal characters (< > : " / \\ | ? *)
    - De‑duplicates with _2, _3, ... suffixes (case‑insensitive)
    """
    stem = Path(path).stem
    name = "".join(ch if ch not in _INVALID_FILE_CHARS else "_" for ch in stem).strip() or "report"
    candidate = f"{name}.txt"
    base = name
    i = 2
    while candidate.lower() in used_lower:
        candidate = f"{base}_{i}.txt"
        i += 1
    used_lower.add(candidate.lower())
    return candidate


# -----------------------------
# Desktop GUI (PySide6)
# -----------------------------

def _run_qt_gui() -> int:  # pragma: no cover (UI not covered by tests)
    try:
        from PySide6.QtWidgets import (
            QApplication,
            QMainWindow,
            QWidget,
            QVBoxLayout,
            QHBoxLayout,
            QPushButton,
            QLabel,
            QFileDialog,
            QMessageBox,
            QTableView,
        )
        from PySide6.QtCore import QAbstractTableModel, Qt, QModelIndex
    except Exception as e:
        print("PySide6 not installed. Install with: pip install PySide6")
        print(f"Details: {e}")
        return 1

    class DataFrameModel(QAbstractTableModel):
        def __init__(self, df: pd.DataFrame | None = None):
            super().__init__()
            self._df = df.astype(str) if df is not None else pd.DataFrame(columns=TARGET_COLS).astype(str)

        def set_df(self, df: pd.DataFrame):
            self.beginResetModel()
            self._df = df.astype(str)
            self.endResetModel()

        def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # type: ignore[override]
            return 0 if self._df is None else len(self._df)

        def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # type: ignore[override]
            return 0 if self._df is None else self._df.shape[1]

        def data(self, index: QModelIndex, role: int = Qt.DisplayRole):  # type: ignore[override]
            if not index.isValid() or role != Qt.DisplayRole:
                return None
            return str(self._df.iat[index.row(), index.column()])

        def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):  # type: ignore[override]
            if role != Qt.DisplayRole:
                return None
            if orientation == Qt.Horizontal:
                return str(self._df.columns[section])
            return str(section + 1)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Process List → TXT Exporter (Desktop)")
            self._df: pd.DataFrame | None = None
            self._dfs: dict[str, pd.DataFrame] = {}  # path -> df

            central = QWidget()
            self.setCentralWidget(central)
            v = QVBoxLayout(central)

            top = QHBoxLayout()
            self.open_btn = QPushButton("Open Excel/CSV (multiple)")
            self.save_btn = QPushButton("Save .txt files (choose names)")
            self.save_btn.setEnabled(False)
            self.status = QLabel("Load a SysGauge report…")
            top.addWidget(self.open_btn)
            top.addWidget(self.save_btn)
            top.addStretch(1)
            top.addWidget(self.status)
            v.addLayout(top)

            self.table = QTableView()
            v.addWidget(self.table)

            self.model = DataFrameModel(pd.DataFrame(columns=TARGET_COLS))
            self.table.setModel(self.model)

            self.open_btn.clicked.connect(self.on_open)
            self.save_btn.clicked.connect(self.on_save)

        def on_open(self):
            paths, _ = QFileDialog.getOpenFileNames(
                self,
                "Open SysGauge reports",
                "",
                "Reports (*.xlsx *.xls *.xlsm *.csv *.csx);;All Files (*)",
            )
            if not paths:
                return
            dfs: dict[str, pd.DataFrame] = {}
            errors: list[str] = []
            for p in paths:
                try:
                    df = extract_process_list_from_path(Path(p))
                    df = reorder_and_format_columns(df)
                    dfs[p] = df
                except Exception as e:
                    errors.append(f"{os.path.basename(p)}: {e}")
            self._dfs = dfs
            if not dfs:
                msg = "No files were processed successfully."
                if errors:
                    msg += "\n" + "\n".join(errors)
                QMessageBox.warning(self, "No valid files", msg)
                self.save_btn.setEnabled(False)
                return
            # Build combined preview with a Source column (from sanitized filename)
            combined: list[pd.DataFrame] = []
            for p, df in dfs.items():
                src = _sheet_title_from_filename(p)
                tmp = df.copy()
                tmp.insert(0, "Source", src)
                combined.append(tmp)
            preview = pd.concat(combined, ignore_index=True) if combined else pd.DataFrame(columns=["Source"] + TARGET_COLS)
            self._df = preview
            self.model.set_df(preview)
            self.save_btn.setEnabled(True)
            total_rows = sum(len(df) for df in dfs.values())
            self.status.setText(f"Loaded {total_rows} rows from {len(dfs)} file(s)")
            if errors:
                QMessageBox.information(self, "Some files were skipped", "\n".join(errors))

        def on_save(self):
            if not self._dfs:
                QMessageBox.information(self, "Nothing to save", "Open one or more files first.")
                return

            written_lower: set[str] = set()
            last_dir = ""
            ok = 0
            skipped = 0
            errors: list[str] = []

            for p, df in self._dfs.items():
                base = os.path.basename(p)
                # propose a safe default name
                default_name = _safe_txt_filename(p, set())  # no de-dupe; just sanitize
                start_path = os.path.join(last_dir or os.getcwd(), default_name)

                sel_path, _ = QFileDialog.getSaveFileName(
                    self,
                    f"Save TXT for {base}",
                    start_path,
                    "Text Files (*.txt);;All Files (*)",
                )
                if not sel_path:
                    skipped += 1
                    continue

                if not sel_path.lower().endswith('.txt'):
                    sel_path += '.txt'

                last_dir = os.path.dirname(sel_path)

                # de-dupe within this export session to avoid accidental overwrite
                if sel_path.lower() in written_lower:
                    root, ext = os.path.splitext(sel_path)
                    i = 2
                    candidate = f"{root}_{i}{ext}"
                    while candidate.lower() in written_lower:
                        i += 1
                        candidate = f"{root}_{i}{ext}"
                    sel_path = candidate

                try:
                    df.to_csv(sel_path, sep="\t", index=False, lineterminator="\n", encoding="utf-8")
                    written_lower.add(sel_path.lower())
                    ok += 1
                except Exception as e:
                    errors.append(f"{base} → {os.path.basename(sel_path)}: {e}")

            msg = f"Saved {ok} .txt file(s)."
            if skipped:
                msg += f"\nSkipped {skipped} file(s) at your request."
            if errors:
                msg += "\n\nSome files failed:\n" + "\n".join(errors)
            QMessageBox.information(self, "Export complete", msg)

    app = QApplication(sys.argv[:1] + [])
    w = MainWindow()
    w.resize(1100, 650)
    w.show()
    app.exec()
    return 0


# -----------------------------
# Tests (no external files required)
# -----------------------------

def _build_sample_csv() -> Tuple[str, str]:
    """Return (filename, csv_text) representing a tiny SysGauge-like CSV."""
    preface = [
        "# SysGauge Process Monitor Report",
        "# Some header garbage we should ignore",
        "",
    ]
    header = ["Process Name,PID,Priority,Threads,CPU,Memory,Handles,Read,Write,Status"]
    rows = [
        "System,4,High,200,0.01,123,10,0,0,Normal",
        "chrome.exe,1111,Normal,42,2.50,500,123,456,789,Normal",
        "notepad.exe,2222,Normal,5,0.00,20,30,0,0,Normal",
        ",,,,,,,,",  # blank tail that should be dropped
    ]
    text = "\n".join(preface + header + rows) + "\n"
    return ("sample.csv", text)


def _build_sample_csv_semicolon() -> Tuple[str, str]:
    """Semicolon-delimited sample to test delimiter handling."""
    preface = ["# Ignore", ""]
    header = ["Process Name;PID;Priority;Threads;CPU;Memory;Status"]
    rows = [
        "System;4;High;200;0.01;123;Normal",
        "explorer.exe;2222;Normal;45;1.5;256;Normal",
    ]
    return ("sample_sc.csv", "\n".join(preface + header + rows) + "\n")


def _build_sample_excel() -> Tuple[str, bytes]:
    """Return (filename, xlsx_bytes) for a tiny Excel with a 'Process List' sheet."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        # A sheet with random summary content
        pd.DataFrame({"A": ["Summary", "More", "Stuff"]}).to_excel(xw, sheet_name="Summary", index=False, header=False)
        # Process List sheet with preface + header + rows
        preface = pd.DataFrame([["SysGauge Report"], ["Ignore this"], [""]])
        header = pd.DataFrame([["Process Name", "PID", "Priority", "Threads", "CPU", "Memory", "Status"]])
        rows = pd.DataFrame([
            ["System", 4, "High", 200, 0.01, 123, "Normal"],
            ["chrome.exe", 1111, "Normal", 42, 2.5, 500, "Normal"],
            ["", "", "", "", "", "", ""],  # blank tail
        ])
        # Write combined to 'Process List'
        preface.to_excel(xw, sheet_name="Process List", index=False, header=False)
        startrow = len(preface)
        header.to_excel(xw, sheet_name="Process List", index=False, header=False, startrow=startrow)
        rows.to_excel(xw, sheet_name="Process List", index=False, header=False, startrow=startrow + 1)
    buf.seek(0)
    return ("sample.xlsx", buf.read())


def run_tests() -> None:
    # CSV test (comma)
    _, text = _build_sample_csv()
    df_csv = extract_process_list_from_filelike(io.StringIO(text), ".csv")
    assert "Process Name" in df_csv.columns[0], "First column should be Process Name"
    assert len(df_csv) == 3, "Should parse 3 non-empty rows"
    assert df_csv.iloc[0][df_csv.columns[0]] == "System"

    # CSV test (semicolon delimiter)
    _, text_sc = _build_sample_csv_semicolon()
    df_csv_sc = extract_process_list_from_filelike(io.StringIO(text_sc), ".csx", csv_delim=";")
    assert df_csv_sc.shape[0] == 2 and df_csv_sc.columns[0].startswith("Process Name")

    # Excel test
    _, xbytes = _build_sample_excel()
    df_xlsx = extract_process_list_from_filelike(io.BytesIO(xbytes), ".xlsx")
    assert df_xlsx.shape[0] == 2, "Should parse 2 non-empty rows in Excel"
    assert df_xlsx.columns[0].startswith("Process Name")

    # Export round-trip test
    out_txt = reorder_and_format_columns(df_xlsx).to_csv(sep="\t", index=False)
    assert out_txt.splitlines()[0].strip() == "\t".join(TARGET_COLS)
    assert "chrome.exe" in out_txt and "System" in out_txt

    # Error case: header not found
    try:
        extract_process_list_from_filelike(io.StringIO("no header here\nfoo,bar\n"), ".csv")
        raise AssertionError("Expected error for missing header")
    except ValueError as e:
        assert "Process Name" in str(e)

    # Column de-duplication test
    cols = _clean_columns(["PID", "PID", "PID", "Process Name"])  # intentional dups
    assert cols[0] == "PID" and cols[1].startswith("PID_") and cols[2].startswith("PID_")

    # Formatting test: thousands separator + 2 decimals + column order
    df_in = pd.DataFrame({
        "Process Name": ["sys", "chrome.exe"],
        "Instances": [1, 17],
        "CPU": [0.1234, 1.725],
        "Memory": [10, 561.777],
        "Threads": [328, 34504],
        "Handles": [6860, 123],
        "Data": [0, 0.654],
        "Status": ["Normal", "Normal"],
    })
    df_fmt = reorder_and_format_columns(df_in)
    assert df_fmt.iloc[1]["Threads"] == "34,504"
    assert df_fmt.iloc[1]["CPU"] == "1.73" and df_fmt.iloc[0]["CPU"] == "0.12"
    assert df_fmt.columns.tolist() == TARGET_COLS

    # Sheet name helpers
    n1 = _sheet_title_from_filename("really:bad/name[with]*chars?.xlsx")
    n2 = _sheet_title_from_filename("really:bad/name[with]*chars? (copy).xlsx")
    used = set()
    u1 = _make_unique_sheet_name(n1, used)
    u2 = _make_unique_sheet_name(n2, used)
    assert u1 != u2 and len(u1) <= 31 and len(u2) <= 31

    # Multi-sheet export smoke test (in-memory)
    from openpyxl import load_workbook
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        used2: set[str] = set()
        for nm in ["alpha_report.xlsx", "alpha_report (copy).xlsx", "bravo.csv"]:
            df_tmp = reorder_and_format_columns(df_in)
            sheet = _make_unique_sheet_name(_sheet_title_from_filename(nm), used2)
            df_tmp.to_excel(xw, sheet_name=sheet, index=False)
    buf.seek(0)
    wb = load_workbook(buf, read_only=True)
    assert len(wb.sheetnames) == 3


# -----------------------------
# Entrypoint
# -----------------------------

def _main() -> int:
    if "--run-tests" in sys.argv:
        run_tests()
        print("All tests passed.")
        return 0
    return _run_qt_gui()


if __name__ == "__main__":
    raise SystemExit(_main())
